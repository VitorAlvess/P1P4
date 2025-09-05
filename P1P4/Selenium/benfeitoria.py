#!/usr/bin/env python
# coding: utf-8

# In[1]:


from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from datetime import datetime
from time import sleep 
import os
import numpy as np
import openpyxl


from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError


# In[2]:


options = webdriver.ChromeOptions()

# options.add_argument('--disable-gpu')
# options.add_argument('--headless')
options.add_argument('--window-size=1366,768')
# options.add_argument('--no-sandbox')
# options.add_argument('--start-maximized')
# options.add_argument('--disable-setuid-sandbox')


preferences = {"download.default_directory" : f'{os.getcwd()}\\relatoriosbenfeitoria', 'profile.default_content_setting_values.automatic_downloads': 1}
options.add_experimental_option("prefs", preferences)
servico = Service(ChromeDriverManager().install())
navegar = webdriver.Chrome(service=servico, options=options)


# In[3]:


pagina = navegar
pagina.get('https://benfeitoria.com/restrito/conta/minhas-campanhas')
sleep(5)


# In[4]:


pagina.find_element('xpath','//*[@id="app"]/main/section/div[1]/div/div[1]/a' ).click()
sleep(5)


# In[5]:


pagina.find_element('xpath','//*[@id="email"]' ).send_keys('pipavoa@opipa.org')
pagina.find_element('xpath','//*[@id="password"]' ).send_keys('Pipavoa21!')
pagina.find_element('xpath','//*[@id="app"]/main/section/div/div[2]/form/div[4]/button' ).click()


# In[6]:


campanhas_ativas = pagina.find_elements('xpath','//*[@id="account"]/div[2]/div/div[2]/div[2]/div/div/div')
print(campanhas_ativas)


# In[41]:


for i in range(len(campanhas_ativas) - 1):
    
    sleep(5)
    pagina.execute_script("window.scrollTo(0, 300);")
    pagina.find_element('xpath',f'//*[@id="account"]/div[2]/div/div[2]/div[2]/div/div/div[{i+1}]' ).click()   
    pagina
    sleep(5)
    pagina.find_element('xpath','//*[@id="project"]/aside/div[1]/ul/li[2]/a' ).click()    
    pagina.find_element('xpath','//*[@id="monitor-container"]/li[5]/a' ).click()
    sleep(5)
    nome_campanha = pagina.find_element('xpath','//*[@id="app"]/header/nav[1]/div[1]/label/p[1]' )
    pagina.find_element('xpath','//*[@id="project"]/div/div[2]/div/div[4]/div/a[1]' ).click()
    pagina.find_element('xpath','//*[@id="project"]/div/div[2]/div/div[4]/div/div/div/div/section[1]/div/form/div/ul/li[1]/label' ).click()
    pagina.find_element('xpath','//*[@id="project"]/div/div[2]/div/div[4]/div/div/div/div/section[2]/div/a' ).click()
    sleep(3)
    # If modifying these scopes, delete the file token.json.
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    creds = None

    if os.path.exists('token.json'): #necessario pegar esse token la da api do google sheets
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())


        # The ID and range of a sample spreadsheet.
    SAMPLE_SPREADSHEET_ID = '1QjpkQAybClkH1Bq9lf-qrNIgJGRB7ZRiHeG07Iae6Ik'
    SAMPLE_RANGE_NAME = 'Campanhas Recorrentes!A:N'


    service = build('sheets', 'v4', credentials=creds)
    #Passar valores existentes para a variavel valore
    sheet = service.spreadsheets()


    result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                    range=SAMPLE_RANGE_NAME).execute()
    valores = result['values']
    # print(valores)
    import os
    import openpyxl

    valores_adicionar = []
    valores_total = []
    valores_final = []
    # Substitua pelo caminho da sua pasta
    pasta = 'relatoriosbenfeitoria/'

    # Lista todos os arquivos na pasta por data de modificação
    arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta) if arquivo.endswith('.xlsx')]
    arquivos = sorted(arquivos, key=lambda x: os.path.getmtime(x), reverse=True)

    # Pega o caminho do último arquivo baixado
    if arquivos:
        ultimo_arquivo = arquivos[0]
        print(f"O último arquivo baixado é: {ultimo_arquivo}")

        # Leitura do arquivo XLSX usando openpyxl
        arquivo_excel = openpyxl.load_workbook(ultimo_arquivo)

        # Escolhe a planilha desejada
        nome_planilha = arquivo_excel.sheetnames[0]  # Pode ajustar conforme necessário
        planilha = arquivo_excel[nome_planilha]

        # Itera sobre as linhas da planilha
        for linha in planilha.iter_rows(min_row=2, values_only=True):
            dados = linha
            valores_adicionar = []
            for elemento in dados:

                valores_adicionar.append(str(elemento))
            valores_adicionar = valores_adicionar[:-7]
            valores_adicionar.append(f'{nome_campanha.text}')
            valores_total.append(valores_adicionar)
    #         print(valores_total)
        # Fecha o arquivo XLSX
        arquivo_excel.close()
    else:
        print("Nenhum arquivo encontrado na pasta.")



    for i in range(len(valores_total)): #Passar os valores que não estão no google sheet
        if valores_total[i] in valores:
            print('O valor está presente')
        else:
             valores_final.append(valores_total[i])
    #         print(valores_total[i])
    #         print('Valores total I acima')

    #         print(valores_final)

    sheet.values().append(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                        range=SAMPLE_RANGE_NAME, valueInputOption="RAW", body={'values': valores_final}).execute()        
#     print(valores)
    print('VALORS FINAIS ABAIXO')
    print(valores_final) #Valores finais.



    sleep(5)
    pagina.get('https://benfeitoria.com/restrito/conta/minhas-campanhas')


# In[43]:


pagina.execute_script("window.scrollTo(0, 300);")
pagina.find_element('xpath',f'//*[@id="account"]/div[2]/div/div[2]/div[2]/div/div/div[{3}]' ).click()   
sleep(2)


# In[44]:


pagina.find_element('xpath','//*[@id="monitor-container"]/li[2]' ).click()   


# In[48]:


nome_campanha = pagina.find_element('xpath','//*[@id="app"]/header/nav[1]/div[1]/label/p[1]' )
pagina.find_element('xpath','//*[@id="project"]/div/div[2]/div[1]/div[3]/div/a[1]' ).click()
sleep(2)
pagina.find_element('xpath','//*[@id="project"]/div/div[2]/div[1]/div[3]/div/div/div/div/section[2]/div/a' ).click()


# In[51]:


SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
creds = None

if os.path.exists('token.json'): #necessario pegar esse token la da api do google sheets
    creds = Credentials.from_authorized_user_file('token.json', SCOPES)
# If there are no (valid) credentials available, let the user log in.
if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            'credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)
    # Save the credentials for the next run
    with open('token.json', 'w') as token:
        token.write(creds.to_json())


    # The ID and range of a sample spreadsheet.
SAMPLE_SPREADSHEET_ID = '1QjpkQAybClkH1Bq9lf-qrNIgJGRB7ZRiHeG07Iae6Ik'
SAMPLE_RANGE_NAME = 'Para Quem Doar!A:N'


service = build('sheets', 'v4', credentials=creds)
#Passar valores existentes para a variavel valore
sheet = service.spreadsheets()


result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                range=SAMPLE_RANGE_NAME).execute()
valores = result['values']
# print(valores)
import os
import openpyxl

valores_adicionar = []
valores_total = []
valores_final = []
# Substitua pelo caminho da sua pasta
pasta = 'relatoriosbenfeitoria/'

# Lista todos os arquivos na pasta por data de modificação
arquivos = [os.path.join(pasta, arquivo) for arquivo in os.listdir(pasta) if arquivo.endswith('.xlsx')]
arquivos = sorted(arquivos, key=lambda x: os.path.getmtime(x), reverse=True)

# Pega o caminho do último arquivo baixado
if arquivos:
    ultimo_arquivo = arquivos[0]
    print(f"O último arquivo baixado é: {ultimo_arquivo}")

    # Leitura do arquivo XLSX usando openpyxl
    arquivo_excel = openpyxl.load_workbook(ultimo_arquivo)

    # Escolhe a planilha desejada
    nome_planilha = arquivo_excel.sheetnames[0]  # Pode ajustar conforme necessário
    planilha = arquivo_excel[nome_planilha]

    # Itera sobre as linhas da planilha
    for linha in planilha.iter_rows(min_row=2, values_only=True):
        dados = linha
        valores_adicionar = []
        for elemento in dados:

            valores_adicionar.append(str(elemento))
        valores_adicionar = valores_adicionar[:-7]
        valores_adicionar.append(f'{nome_campanha.text}')
        valores_total.append(valores_adicionar)
#         print(valores_total)
    # Fecha o arquivo XLSX
    arquivo_excel.close()
else:
    print("Nenhum arquivo encontrado na pasta.")



for i in range(len(valores_total)): #Passar os valores que não estão no google sheet
    if valores_total[i] in valores:
        print('O valor está presente')
    else:
         valores_final.append(valores_total[i])
#         print(valores_total[i])
#         print('Valores total I acima')

#         print(valores_final)

sheet.values().append(spreadsheetId=SAMPLE_SPREADSHEET_ID,
                                    range=SAMPLE_RANGE_NAME, valueInputOption="RAW", body={'values': valores_final}).execute()        
#     print(valores)
print('VALORS FINAIS ABAIXO')
print(valores_final) #Valores finais.



sleep(5)
pagina.get('https://benfeitoria.com/restrito/conta/minhas-campanhas')


# In[ ]:




